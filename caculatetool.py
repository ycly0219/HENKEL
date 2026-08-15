# -*- coding: utf-8 -*-
"""
调拨计算工具 - WMS库存查询 GUI 小程序

功能：
1. 导入调拨计算 Excel（含 SKU / 库存地 / 质量状态 / 箱数 / warehouseId 列），表格预览数据
2. 逐行调用 WMS 库存查询 API，按 Excel 中的 warehouseId 动态查询库存明细
3. 明细表展示：WAREHOUSEID / 箱规 / 散支数量(qty) / 失效日期 / 实物批次 / 质量状态 / 库存地 / SAP工厂(lotAtt14)
   - 散支(qty < 箱规)筛选：整箱记录(qty>=箱规)按失效日期升序累加至 >= 需求箱数×箱规，
     以最后一条整箱记录的失效日期为「截止日期」；仅展示失效日期 <= 截止日期的散支记录
   - 接口无返回明细 → 标注"无库存"；无散支候选 → 标注"无散支"；散支均晚于截止日期 → 标注"无符合截止期散支"
4. 汇总：总调拨数 = 需求箱数 × 箱规；仅当存在符合截止期的散支时，再加筛选后散支数量之和
   - 汇总表首列展示 WAREHOUSEID，并新增"含散支"列（含散支=Y 整行红色且置顶优先）
5. 仓库ID校验：单张Excel只允许包含 WH004004 或 WH004013 其中之一，二者并存则拒绝导入
6. 导出 Excel：在原表基础上新增"箱规/整箱支数/含散支/总调拨数"四列，文件名带时间戳

依赖：openpyxl、requests（pip install openpyxl requests）
"""

import json
import os
import re
import threading
import traceback
from datetime import datetime, date, timedelta

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    import requests
except ImportError:
    requests = None

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    openpyxl = None

from template_data import get_template_bytes

# api中将组织号更换为实际组织号
# ============================ 配置 ============================
API_URL = ("https://sinoewms.i./组织号/.com/datahubjson/"
           "FLUXWMSUDF_C231_HD01/JsonApi/?method=QUERYINV")
TIMEOUT_SECONDS = 15          # 单次API请求超时时间(秒)
# 注：warehouseId 不再使用固定值，改为从 Excel 的 warehouseId 列动态读取

# Excel 列名候选（不区分大小写、忽略首尾空格）
COL_ALIASES = {
    "sku":    ["sku", "SKU", "商品编码", "物料编码"],
    "name":   ["name", "品名", "商品名称", "名称"],
    "loc":    ["库存地", "lotatt10", "库存地点"],
    "status": ["质量状态", "lotatt08", "状态"],
    "boxes":  ["箱数", "需求箱数", "调拨箱数", "box", "boxes"],
    "wid":    ["warehouseid", "仓库id", "仓库编码", "仓库代码",
               "库存仓库", "物流仓库", "仓库", "wh"],
}


# ============================ 工具函数 ============================
def norm(s):
    """列名规范化：去空格转小写"""
    return str(s).strip().lower() if s is not None else ""


def to_num(v, default=0):
    """安全转数字"""
    if v is None or v == "":
        return default
    try:
        f = float(str(v).strip())
        return int(f) if f == int(f) else f
    except (ValueError, TypeError):
        return default


def fmt_num(v):
    """数字显示格式化"""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v)


def note_rank(note):
    """
    库存明细排序优先级（数值越小越靠前）：
      0 - 无库存
      1 - 无散支
      2 - 其他有备注（如查询失败 / 接口未回告箱规 等错误提示）
      3 - 无备注（正常散支明细，排最后）
    """
    if note == "无库存":
        return 0
    if note == "无散支":
        return 1
    if note:                       # 有备注但非上述两类（错误信息）
        return 2
    return 3                       # 无备注，排最后


# ============================ 日期解析（散支筛选用） ============================
def parse_expiry(s):
    """解析失效日期(lotAtt02)为 date；失败返回 None。
    支持 YYYY-MM-DD / YYYY/MM/DD / YYYYMMDD 及带时分秒的形式，
    也兼容 Excel 序列号（整数）。"""
    if s is None:
        return None
    if isinstance(s, datetime):
        return s.date()
    if isinstance(s, date):
        return s
    t = str(s).strip()
    if not t:
        return None
    # Excel 序列号（数字）
    if re.fullmatch(r"\d+(\.\d+)?", t):
        try:
            serial = float(t)
            if 1 <= serial < 100000:
                return date(1899, 12, 30) + timedelta(days=int(serial))
        except (ValueError, OverflowError, OSError):
            pass
    fmts = ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d",
            "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S",
            "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M")
    for f in fmts:
        try:
            return datetime.strptime(t, f).date()
        except ValueError:
            continue
    return None


def expiry_sort_key(s):
    """排序键：失效日期可解析的在前(升序)，不可解析的排最后。"""
    d = parse_expiry(s)
    if d is None:
        return (1, date(9999, 12, 31))
    return (0, d)


def compute_loose_filter(records, box_spec, boxes):
    """按 FEFO 规则筛选散支并计算截止日期。

    规则：
      1. 整箱记录(qty >= box_spec)按失效日期升序，依次累加 qty 直至
         累加总和 >= 总需求数量(需求箱数 × 箱规)，记录最后参与累加的
         整箱记录的失效日期作为「截止日期」。
      2. 散支记录(qty < box_spec)中，失效日期 <= 截止日期 的全部筛选出。

    返回 (filtered_loose, cutoff_raw, cutoff_date, loose_sum)：
      filtered_loose : 通过日期筛选的散支记录列表
      cutoff_raw     : 截止日期原始字符串（无则为 None）
      cutoff_date    : 解析后的截止日期（date 或 None）
      loose_sum      : 筛选后散支 qty 之和
    """
    demand = boxes * box_spec

    # 整箱记录按失效日期升序
    fullbox = [r for r in records if to_num(r.get("qty"), 0) >= box_spec]
    fullbox.sort(key=lambda r: expiry_sort_key(r.get("lotAtt02")))

    cutoff_raw = None
    if demand > 0:
        cum = 0
        for r in fullbox:
            cum += to_num(r.get("qty"), 0)
            cutoff_raw = r.get("lotAtt02")
            if cum >= demand:
                break
    cutoff_date = parse_expiry(cutoff_raw)

    # 散支候选 + 日期筛选
    loose_cands = [r for r in records if to_num(r.get("qty"), 0) < box_spec]
    filtered = []
    if cutoff_date is not None:
        for r in loose_cands:
            d = parse_expiry(r.get("lotAtt02"))
            if d is not None and d <= cutoff_date:
                filtered.append(r)

    loose_sum = sum(to_num(r.get("qty"), 0) for r in filtered)
    return filtered, cutoff_raw, cutoff_date, loose_sum


# 允许的仓库ID：单张 Excel 只允许包含其中之一
ALLOWED_WAREHOUSES = {"WH004004", "WH004013"}


def validate_warehouses(rows):
    """校验单张 Excel 只允许包含 WH004004 或 WH004013 其中之一，二者不可并存。
    返回错误信息字符串；校验通过返回 None。"""
    wids = set()
    for r in rows:
        w = str(r.get("wid") or "").strip().upper()
        if w:
            wids.add(w)
    if not wids:
        return "未检测到任何有效的仓库ID（warehouseId 列应为 WH004004 或 WH004013）"
    present = wids & ALLOWED_WAREHOUSES
    if len(present) > 1:
        return "仓库ID不合规：同一 Excel 不允许同时包含 WH004004 与 WH004013，请拆分为两个文件分别导入"
    bad = wids - ALLOWED_WAREHOUSES
    if bad:
        return "仓库ID不合规：只允许 WH004004 或 WH004013，检测到非法仓库ID：%s" % "、".join(sorted(bad))
    return None


# ============================ API 调用 ============================
def query_inventory(sku, lot_att08, lot_att10, warehouse_id):
    """
    调用 WMS 库存查询接口。
    返回 (记录列表, 错误信息)。成功时错误信息为 None。
    每条记录: {userDefine1, qty, lotAtt02, lotAtt04, lotAtt08, lotAtt10, lotAtt14}
    """
    if requests is None:
        return [], "缺少 requests 库，请先执行: pip install requests"

    payload = {
        "data": {
            "header": {
                "warehouseId": str(warehouse_id).strip(),
                "sku": str(sku).strip(),
                "lotAtt08": str(lot_att08).strip(),
                "lotAtt10": str(lot_att10).strip(),
            }
        }
    }
    try:
        resp = requests.post(
            API_URL,
            json=payload,
            headers={"Content-Type": "application/json"},
            timeout=TIMEOUT_SECONDS,
            verify=False,
        )
    except requests.exceptions.Timeout:
        return [], "请求超时(%ds)" % TIMEOUT_SECONDS
    except requests.exceptions.ConnectionError as e:
        return [], "网络连接失败: %s" % str(e)[:120]
    except Exception as e:
        return [], "请求异常: %s" % str(e)[:120]

    if resp.status_code != 200:
        return [], "HTTP %d" % resp.status_code

    try:
        data = resp.json()
    except ValueError:
        return [], "响应不是有效JSON: %s" % resp.text[:120]

    # 递归收集所有含 qty 或 userDefine1 字段的明细 dict（兼容不同层级的返回结构）
    records = []

    def collect(node):
        if isinstance(node, dict):
            keys = {k.lower() for k in node.keys()}
            if "qty" in keys or "userdefine1" in keys:
                low = {k.lower(): v for k, v in node.items()}
                records.append({
                    "userDefine1": low.get("userdefine1", ""),
                    "qty":         low.get("qty", 0),
                    "lotAtt02":    low.get("lotatt02", ""),
                    "lotAtt04":    low.get("lotatt04", ""),
                    "lotAtt08":    low.get("lotatt08", ""),
                    "lotAtt10":    low.get("lotatt10", ""),
                    "lotAtt14":    low.get("lotatt14", ""),
                })
            else:
                for v in node.values():
                    collect(v)
        elif isinstance(node, list):
            for item in node:
                collect(item)

    collect(data)

    if not records:
        # 尝试提取接口返回的提示信息
        msg = ""
        if isinstance(data, dict):
            for key in ("message", "msg", "errorMsg", "returnDesc", "desc"):
                for k, v in data.items():
                    if k.lower() == key.lower() and v:
                        msg = str(v)
                        break
                if msg:
                    break
        return [], ("无库存记录" + ("（%s）" % msg if msg else ""))

    return records, None


# ============================ 主程序 ============================
class TransferApp:
    def __init__(self, root):
        self.root = root
        root.title("调拨计算工具 - WMS库存查询")
        root.geometry("1180x720")
        root.minsize(920, 600)

        self.excel_path = None      # 导入的Excel路径
        self.excel_rows = []        # [{sku,name,loc,status,boxes}, ...]
        self.detail_rows = []       # 查询后的明细行
        self.summary = {}           # sku -> {box_spec, qty_sum, boxes, total, error}
        self.querying = False

        self._build_ui()

    # ---------- UI ----------
    def _build_ui(self):
        style = ttk.Style()
        try:
            style.theme_use("vista")
        except tk.TclError:
            pass
        style.configure("Treeview", rowheight=24)
        style.configure("Treeview.Heading", font=("Microsoft YaHei UI", 9, "bold"))

        default_font = ("Microsoft YaHei UI", 9)

        # 顶部按钮栏
        top = ttk.Frame(self.root, padding=(10, 8))
        top.pack(fill="x")

        self.btn_import = ttk.Button(top, text="📂 导入Excel", command=self.on_import)
        self.btn_import.pack(side="left")

        self.btn_template = ttk.Button(top, text="📥 导入模板下载",
                                       command=self.on_download_template)
        self.btn_template.pack(side="left", padx=(8, 0))

        self.btn_query = ttk.Button(top, text="🔍 查询库存并计算", command=self.on_query,
                                    state="disabled")
        self.btn_query.pack(side="left", padx=(8, 0))

        self.btn_export = ttk.Button(top, text="💾 导出Excel", command=self.on_export,
                                     state="disabled")
        self.btn_export.pack(side="left", padx=(8, 0))

        self.lbl_file = ttk.Label(top, text="未导入文件", foreground="#666",
                                  font=default_font)
        self.lbl_file.pack(side="left", padx=(16, 0))

        # 进度条
        self.progress = ttk.Progressbar(top, length=180, mode="determinate")
        self.progress.pack(side="right")

        # 主区域：Notebook 两个页签
        nb = ttk.Notebook(self.root)
        nb.pack(fill="both", expand=True, padx=10, pady=(0, 4))
        self.nb = nb

        # -- 页签1：Excel预览 --
        tab1 = ttk.Frame(nb)
        nb.add(tab1, text="  Excel数据预览  ")
        cols1 = ("wid", "sku", "name", "loc", "status", "boxes")
        heads1 = ("WAREHOUSEID", "SKU", "品名", "库存地", "质量状态", "需求箱数")
        widths1 = (110, 100, 380, 90, 90, 80)
        self.tv_excel = self._make_tree(tab1, cols1, heads1, widths1)

        # -- 页签2：库存明细 --
        tab2 = ttk.Frame(nb)
        nb.add(tab2, text="  库存明细（API回告） ")
        cols2 = ("wid", "sku", "name", "box_spec", "qty", "lot02", "lot04",
                 "lot08", "lot10", "lot14", "note")
        heads2 = ("WAREHOUSEID", "SKU", "品名", "箱规", "散支数量(qty)", "失效日期",
                  "实物批次", "质量状态", "库存地", "SAP工厂(lotAtt14)", "备注")
        widths2 = (110, 100, 280, 80, 110, 110, 130, 90, 90, 110, 200)
        self.tv_detail = self._make_tree(tab2, cols2, heads2, widths2)
        self.tv_detail.tag_configure("error", foreground="#c62828")
        self.tv_detail.tag_configure("info", foreground="#1565c0")
        self.tv_detail.tag_configure("odd", background="#f5f7fa")

        # -- 底部汇总区 --
        sum_frame = ttk.LabelFrame(self.root, text=" 汇总：总调拨数 = 需求箱数 × 箱规 ＋ 筛选后散支数量(仅当含符合截止期的散支时)；含散支行整行红色、置顶优先 ",
                                   padding=(8, 4))
        sum_frame.pack(fill="both", padx=10, pady=(2, 6), ipady=2)

        cols3 = ("wid", "sku", "name", "boxes", "box_spec", "box_total", "loose_cnt", "loose", "total")
        heads3 = ("WAREHOUSEID", "SKU", "品名", "需求箱数", "箱规", "整箱支数", "散支支数", "含散支", "总调拨数")
        widths3 = (110, 100, 330, 80, 80, 110, 90, 80, 130)
        self.tv_sum = self._make_tree(sum_frame, cols3, heads3, widths3, height=6)
        self.tv_sum.tag_configure("error", foreground="#c62828")
        self.tv_sum.tag_configure("redrow", foreground="#c62828")
        self.tv_sum.tag_configure("totalrow", background="#fff3cd",
                                  font=("Microsoft YaHei UI", 9, "bold"))

        # 状态栏
        self.status_var = tk.StringVar(value="请先导入调拨计算Excel文件")
        bar = ttk.Frame(self.root)
        bar.pack(fill="x", side="bottom")
        ttk.Label(bar, textvariable=self.status_var, foreground="#444",
                  font=default_font, padding=(10, 3)).pack(side="left")

    def _make_tree(self, parent, cols, heads, widths, height=None):
        """创建带横/纵滚动条的 Treeview"""
        frame = ttk.Frame(parent)
        frame.pack(fill="both", expand=True)

        kwargs = {"columns": cols, "show": "headings"}
        if height:
            kwargs["height"] = height
        tv = ttk.Treeview(frame, **kwargs)
        for c, h, w in zip(cols, heads, widths):
            tv.heading(c, text=h, anchor="center")
            tv.column(c, width=w, minwidth=60, anchor="center", stretch=False)

        vsb = ttk.Scrollbar(frame, orient="vertical", command=tv.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=tv.xview)
        tv.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tv.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)
        return tv

    # ---------- 导入模板下载 ----------
    def on_download_template(self):
        """将内嵌的调拨计算模板另存为用户选择的 xlsx 文件。"""
        if self.querying:
            return
        path = filedialog.asksaveasfilename(
            title="保存导入模板",
            defaultextension=".xlsx",
            initialfile="调拨计算导入模板.xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
        )
        if not path:
            return
        try:
            with open(path, "wb") as f:
                f.write(get_template_bytes())
        except Exception as e:
            messagebox.showerror("保存失败", "无法保存导入模板：\n%s" % e)
            return
        self.status_var.set("已保存导入模板：" + path)
        messagebox.showinfo("保存成功", "导入模板已保存到：\n%s" % path)

    # ---------- 导入 ----------
    def on_import(self):
        if self.querying:
            return
        path = filedialog.askopenfilename(
            title="选择调拨计算Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xlsm"), ("所有文件", "*.*")],
        )
        if not path:
            return
        if openpyxl is None:
            messagebox.showerror("缺少依赖", "缺少 openpyxl 库，请先执行:\npip install openpyxl")
            return
        try:
            rows = self._load_excel(path)
        except Exception as e:
            messagebox.showerror("读取失败", "无法读取Excel文件：\n%s" % e)
            return
        werr = validate_warehouses(rows)
        if werr:
            messagebox.showerror("仓库ID校验失败", werr)
            return
        if not rows:
            messagebox.showwarning("无数据", "未在Excel中找到有效数据行，请检查文件内容与列名\n"
                                   "（需要包含：SKU、库存地、质量状态、箱数、warehouseId仓库ID）")
            return

        self.excel_path = path
        self.excel_rows = rows
        self.detail_rows = []
        self.summary = {}

        # 刷新预览表
        self.tv_excel.delete(*self.tv_excel.get_children())
        for i, r in enumerate(rows):
            tag = ("odd",) if i % 2 else ()
            self.tv_excel.insert("", "end", values=(
                r["wid"], r["sku"], r["name"], r["loc"], r["status"],
                fmt_num(r["boxes"])), tags=tag)
        self.tv_detail.delete(*self.tv_detail.get_children())
        self.tv_sum.delete(*self.tv_sum.get_children())

        self.lbl_file.config(text=os.path.basename(path))
        self.btn_query.config(state="normal")
        self.btn_export.config(state="disabled")
        self.nb.select(0)
        self.status_var.set("已导入 %d 行数据，点击【查询库存并计算】开始" % len(rows))

    def _load_excel(self, path):
        """读取Excel，自动识别列名，返回行列表"""
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active

        # 定位表头行（前5行内查找包含 sku 的行）
        header_row_idx, col_map = None, {}
        for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
            cells = [norm(c) for c in row]
            if any(c == "sku" or "sku" in c for c in cells if c):
                header_row_idx = ridx
                for cidx, cell in enumerate(cells):
                    for field, aliases in COL_ALIASES.items():
                        if cell and cell in [norm(a) for a in aliases]:
                            col_map.setdefault(field, cidx)
                break
        if header_row_idx is None:
            raise ValueError("未找到表头行（需包含 SKU 列）")
        for need in ("sku", "loc", "status", "boxes", "wid"):
            if need not in col_map:
                raise ValueError("缺少必需列：%s（当前识别到 %s）" %
                                 ({"sku": "SKU", "loc": "库存地",
                                   "status": "质量状态", "boxes": "箱数",
                                   "wid": "warehouseId(仓库ID)"}[need],
                                  list(col_map.keys())))

        rows = []
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            sku = row[col_map["sku"]] if col_map["sku"] < len(row) else None
            if sku is None or str(sku).strip() == "":
                continue
            get = lambda f: (row[col_map[f]] if f in col_map and col_map[f] < len(row) else "")
            sku_s = str(sku).strip()
            if sku_s.endswith(".0"):     # 数字SKU去掉小数尾巴
                sku_s = sku_s[:-2]
            rows.append({
                "sku": sku_s,
                "name": str(get("name") or "").strip(),
                "loc": str(get("loc") or "").strip(),
                "status": str(get("status") or "").strip(),
                "boxes": to_num(get("boxes"), 0),
                "wid": str(get("wid") or "").strip(),
            })
        return rows

    # ---------- 查询 ----------
    def on_query(self):
        if self.querying or not self.excel_rows:
            return
        self.querying = True
        self.btn_import.config(state="disabled")
        self.btn_query.config(state="disabled")
        self.btn_export.config(state="disabled")
        self.tv_detail.delete(*self.tv_detail.get_children())
        self.tv_sum.delete(*self.tv_sum.get_children())
        self.progress.config(maximum=len(self.excel_rows), value=0)
        self.nb.select(1)
        self.status_var.set("正在查询库存...")

        threading.Thread(target=self._query_worker, daemon=True).start()

    def _query_worker(self):
        """后台线程：逐行调用API"""
        # 屏蔽自签名证书告警
        try:
            import urllib3
            urllib3.disable_warnings()
        except Exception:
            pass

        details, summary = [], {}
        total_n = len(self.excel_rows)
        for idx, r in enumerate(self.excel_rows, 1):
            sku = r["sku"]
            self.root.after(0, self._on_progress, idx, total_n, sku)

            records, err = query_inventory(sku, r["status"], r["loc"], r["wid"])

            if err:
                details.append({"wid": r["wid"], "sku": sku, "name": r["name"], "box_spec": "",
                                "qty": "", "lot02": "", "lot04": "",
                                "lot08": "", "lot10": "", "lot14": "",
                                "note": "查询失败: " + err, "kind": "error"})
                summary[sku] = {"wid": r["wid"], "name": r["name"], "boxes": r["boxes"],
                                "box_spec": None, "total": None,
                                "has_loose": False, "error": err, "kind": "error"}
            elif not records:
                details.append({"wid": r["wid"], "sku": sku, "name": r["name"], "box_spec": "",
                                "qty": "", "lot02": "", "lot04": "",
                                "lot08": "", "lot10": "", "lot14": "",
                                "note": "无库存", "kind": "info"})
                summary[sku] = {"wid": r["wid"], "name": r["name"], "boxes": r["boxes"],
                                "box_spec": None, "total": None,
                                "has_loose": False, "error": "无库存", "kind": "info"}
            else:
                # 取该SKU箱规（首个非空的 userDefine1）
                box_spec = None
                for rec in records:
                    spec = to_num(rec.get("userDefine1"), None)
                    if box_spec is None and spec:
                        box_spec = spec
                if box_spec is None:
                    details.append({"wid": r["wid"], "sku": sku, "name": r["name"], "box_spec": "",
                                    "qty": "", "lot02": "", "lot04": "",
                                    "lot08": "", "lot10": "", "lot14": "",
                                    "note": "接口未回告箱规(userDefine1)", "kind": "error"})
                    summary[sku] = {"wid": r["wid"], "name": r["name"], "boxes": r["boxes"],
                                    "box_spec": None, "total": None,
                                    "has_loose": False,
                                    "error": "接口未回告箱规(userDefine1)", "kind": "error"}
                else:
                    # 新散支筛选逻辑（FEFO）：
                    #   1) 整箱记录(qty>=箱规)按失效日期升序累加至 >= 需求箱数×箱规，
                    #      以最后一条整箱记录的失效日期为截止日期
                    #   2) 散支记录(qty<箱规)中，失效日期 <= 截止日期 的全部筛选出
                    filtered, cutoff_raw, cutoff_date, loose_sum = \
                        compute_loose_filter(records, box_spec, r["boxes"])

                    if not filtered:
                        # 无符合截止日期的散支：本就无 qty<箱规 的散支，
                        # 或散支失效日期均晚于截止日期 → 明细不展示，按纯整箱计算
                        has_loose_cand = any(to_num(rec.get("qty"), 0) < box_spec
                                             for rec in records)
                        note = ("无散支" if not has_loose_cand
                                else "无符合截止期散支(截止:%s)"
                                      % (str(cutoff_raw) if cutoff_raw else "无"))
                        details.append({"wid": r["wid"], "sku": sku, "name": r["name"],
                                        "box_spec": fmt_num(box_spec),
                                        "qty": "", "lot02": "", "lot04": "",
                                        "lot08": "", "lot10": "", "lot14": "",
                                        "note": note, "kind": "info"})
                        total = r["boxes"] * box_spec
                        summary[sku] = {"wid": r["wid"], "name": r["name"],
                                        "boxes": r["boxes"],
                                        "box_spec": box_spec, "total": total,
                                        "loose_sum": 0,
                                        "has_loose": False, "kind": "valid"}
                    else:
                        for rec in filtered:
                            q = to_num(rec["qty"], 0)
                            details.append({"wid": r["wid"], "sku": sku, "name": r["name"],
                                            "box_spec": fmt_num(box_spec),
                                            "qty": fmt_num(q),
                                            "lot02": str(rec.get("lotAtt02") or ""),
                                            "lot04": str(rec.get("lotAtt04") or ""),
                                            "lot08": str(rec.get("lotAtt08") or ""),
                                            "lot10": str(rec.get("lotAtt10") or ""),
                                            "lot14": str(rec.get("lotAtt14") or ""),
                                            "note": "", "kind": "normal"})
                        # 含散支：总调拨数 = 需求箱数 × 箱规 + 筛选后散支数量之和
                        total = r["boxes"] * box_spec + loose_sum
                        summary[sku] = {"wid": r["wid"], "name": r["name"],
                                        "boxes": r["boxes"],
                                        "box_spec": box_spec, "total": total,
                                        "loose_sum": loose_sum,
                                        "has_loose": True, "kind": "valid"}

        self.detail_rows = details
        # 库存明细按备注排序：无库存 > 无散支 > 错误提示 > 无备注(正常散支,最后)
        self.detail_rows.sort(key=lambda d: (note_rank(d.get("note", "")),
                                             d["sku"],
                                             expiry_sort_key(d.get("lot02"))))
        self.summary = summary
        self.root.after(0, self._on_query_done)

    def _on_progress(self, idx, total, sku):
        self.progress.config(value=idx)
        self.status_var.set("正在查询 %d/%d：SKU %s ..." % (idx, total, sku))

    def _on_query_done(self):
        # 明细表
        self.tv_detail.delete(*self.tv_detail.get_children())
        for i, d in enumerate(self.detail_rows):
            # 只展示散支记录（kind=normal）；无散支 / 无库存 / 查询失败 均不展示
            if d["kind"] != "normal":
                continue
            tags = []
            if i % 2:
                tags.append("odd")
            self.tv_detail.insert("", "end", values=(
                d["wid"], d["sku"], d["name"], d["box_spec"], d["qty"], d["lot02"],
                d["lot04"], d["lot08"], d["lot10"], d["lot14"], d["note"]), tags=tuple(tags))

        # 汇总表（含散支行优先排前，并以红色标示）
        self.tv_sum.delete(*self.tv_sum.get_children())
        grand_total = 0
        has_valid = False
        fail_cnt = 0
        loose_cnt = 0

        # 收集并按"是否含散支"分组排序：含散支(Y)的行排在前面
        entries = [(r, self.summary[r["sku"]])
                   for r in self.excel_rows if r["sku"] in self.summary]
        entries.sort(key=lambda e: (0 if e[1].get("has_loose") else 1,))

        for r, s in entries:
            k = s.get("kind")
            wid = s.get("wid", "")
            loose = "Y" if s.get("has_loose") else "N"
            if loose == "Y":
                loose_cnt += 1
            if k == "error":
                fail_cnt += 1
                self.tv_sum.insert("", "end", values=(
                    wid, r["sku"], s["name"], fmt_num(s["boxes"]), "-", "-",
                    "-", loose, s["error"]), tags=("error",))
            elif k == "info":
                # 无库存等情况，仅提示、不计入合计
                self.tv_sum.insert("", "end", values=(
                    wid, r["sku"], s["name"], fmt_num(s["boxes"]), "-", "-",
                    "-", loose, s["error"]), tags=("error",))
            else:
                has_valid = True
                grand_total += s["total"]
                tags = ("redrow",) if loose == "Y" else ()
                self.tv_sum.insert("", "end", values=(
                    wid, r["sku"], s["name"], fmt_num(s["boxes"]),
                    fmt_num(s["box_spec"]), fmt_num(s["boxes"] * s["box_spec"]),
                    fmt_num(s.get("loose_sum", 0)), loose, fmt_num(s["total"])), tags=tags)
        # 全局合计行
        self.tv_sum.insert("", "end", values=(
            "", "合计", "", "", "", "", "", fmt_num(grand_total)), tags=("totalrow",))

        self.querying = False
        self.btn_import.config(state="normal")
        self.btn_query.config(state="normal")
        self.btn_export.config(state="normal" if has_valid or self.summary else "disabled")

        msg = "查询完成：共 %d 个SKU" % len(self.summary)
        if fail_cnt:
            msg += "，其中 %d 个失败（明细表中红色标注）" % fail_cnt
            messagebox.showwarning("部分查询失败",
                                   "%d 个SKU查询失败，请查看明细表红色行的错误信息。" % fail_cnt)
        msg += "，全局总调拨数 = %s" % fmt_num(grand_total)
        self.status_var.set(msg)

    # ---------- 导出 ----------
    def on_export(self):
        if not self.summary or not self.excel_path:
            messagebox.showwarning("无数据", "请先导入Excel并完成查询计算")
            return
        try:
            out_path = self._export_excel()
        except Exception as e:
            # 记录完整堆栈，便于定位导出异常
            err_path = os.path.join(os.path.dirname(__file__), "_export_error.log")
            with open(err_path, "w", encoding="utf-8") as f:
                traceback.print_exc(file=f)
            messagebox.showerror("导出失败", "导出Excel时出错：\n%s\n\n详细堆栈已保存到：\n%s" % (e, err_path))
            return
        self.status_var.set("已导出：" + out_path)
        if messagebox.askyesno("导出成功", "文件已导出到：\n%s\n\n是否打开所在文件夹？" % out_path):
            try:
                os.startfile(os.path.dirname(out_path))
            except Exception:
                pass

    def _export_excel(self):
        """在原Excel基础上新增'总调拨数'列，另存为带时间戳的新文件"""
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active

        # 找表头行和SKU列
        header_row_idx, sku_col = None, None
        for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=5), 1):
            for cell in row:
                if norm(cell.value) == "sku" or (cell.value and "sku" in norm(cell.value)):
                    header_row_idx, sku_col = ridx, cell.column
                    break
            if header_row_idx:
                break
        if header_row_idx is None:
            raise ValueError("原Excel中未找到SKU表头")

        # 新增导出列：箱规 / 整箱支数 / 含散支 / 散支数量 / 总调拨数（字段名与系统汇总表一致）
        export_cols = ["箱规", "整箱支数", "含散支", "散支数量", "总调拨数"]
        start_col = ws.max_column + 1
        col_indices = []
        for i, name in enumerate(export_cols):
            c = start_col + i
            col_indices.append(c)
            hc = ws.cell(row=header_row_idx, column=c, value=name)
            hc.font = Font(bold=True)
            hc.alignment = Alignment(horizontal="center")
            hc.fill = PatternFill("solid", fgColor="FFF2CC")
            ws.column_dimensions[hc.column_letter].width = 12

        for row in ws.iter_rows(min_row=header_row_idx + 1):
            sku_val = row[sku_col - 1].value
            if sku_val is None or str(sku_val).strip() == "":
                continue
            sku_s = str(sku_val).strip()
            if sku_s.endswith(".0"):
                sku_s = sku_s[:-2]
            s = self.summary.get(sku_s)
            if s is None:
                continue
            err = s.get("error")
            ridx = row[0].row
            if err:
                ws.cell(row=ridx, column=col_indices[0], value="")   # 箱规
                ws.cell(row=ridx, column=col_indices[1], value="")   # 整箱支数
                ws.cell(row=ridx, column=col_indices[2],
                        value="Y" if s.get("has_loose") else "N")    # 含散支
                ws.cell(row=ridx, column=col_indices[3],
                        value=(fmt_num(s.get("loose_sum", 0)) if s.get("loose_sum") else ""))  # 散支数量
                ec = ws.cell(row=ridx, column=col_indices[4], value=str(err))  # 总调拨数
                ec.font = Font(color="C62828")
            else:
                box_spec = s.get("box_spec")
                boxes = s.get("boxes")
                box_total = (boxes * box_spec) if (box_spec is not None and boxes is not None) else None
                total_val = s.get("total")
                loose_val = s.get("loose_sum", 0)
                ws.cell(row=ridx, column=col_indices[0],
                        value=(fmt_num(box_spec) if box_spec is not None else ""))
                ws.cell(row=ridx, column=col_indices[1],
                        value=(fmt_num(box_total) if box_total is not None else ""))
                ws.cell(row=ridx, column=col_indices[2],
                        value="Y" if s.get("has_loose") else "N")
                ws.cell(row=ridx, column=col_indices[3],
                        value=(fmt_num(loose_val) if loose_val else ""))
                ws.cell(row=ridx, column=col_indices[4],
                        value=(total_val if total_val is not None else ""))

        base, _ = os.path.splitext(os.path.basename(self.excel_path))
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join(os.path.dirname(self.excel_path),
                                "%s_总调拨数_%s.xlsx" % (base, ts))
        wb.save(out_path)
        return out_path


# ============================ 入口 ============================
def main():
    missing = []
    if openpyxl is None:
        missing.append("openpyxl")
    if requests is None:
        missing.append("requests")
    root = tk.Tk()
    app = TransferApp(root)
    if missing:
        messagebox.showwarning(
            "缺少依赖",
            "缺少以下依赖库：%s\n请先执行：pip install %s" %
            (", ".join(missing), " ".join(missing)))
    root.mainloop()


if __name__ == "__main__":
    main()
